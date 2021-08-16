#!/usr/bin/python

import io
import json
import openpyxl
import struct
import sys
import zipfile

def safer(values, key):
    if values.has_key(key):
        return values[key]
    else:
        return None

def real(value):
    if value == None:
        return None
    else:
        sign = value['sign']
        mantissa = value['mantissa']
        exponent = value['exponent']
        value = (sign << 31) | (exponent << 23) | mantissa
        return struct.unpack('<f', struct.pack('<L', value))[0]

type_to_format = {
    -1: '<q',
    0: '<qfff',
    1: '<qfff',
    2: '<qfff',
    3: '<qfff',
    4: '<qfff',
    5: '<qf',
    6: '<qf',
    7: '<qf',
    8: '<qf',
    9: '<qfff',
    10: '<qfff',
    11: '<qfff',
    12: '<qf',
    13: '<qf',
    14: '<qffffff',
    15: '<qfff',
    16: '<qffffff',
    17: '<q',
    18: '<q',
    19: '<qf',
    20: '<qfff',
    21: '<qfff'}

def device(device_json, device_sheet):
    device_sheet.title = 'Device'
    device_sheet.append(['ID', device_json['device']['id']])
    device_sheet.append(['Brand', device_json['device']['brand']])
    device_sheet.append(['Manufacturer', device_json['device']['manufacturer']])
    device_sheet.append(['Product', device_json['device']['product']])
    device_sheet.append(['Model', device_json['device']['model']])
    device_sheet.append(['Design', device_json['device']['design']])
    device_sheet.append(['Hardware', device_json['device']['hardware']])
    device_sheet.append(['Board', device_json['device']['board']])
    device_sheet.append(['Build', device_json['device']['build']])

def sensors(device_json, sensors_sheet):
    sensors_sheet.title = 'Sensors'
    sensors_sheet.append(['Index', 'Type', 'Name', 'Vendor', 'Delay', 'Resolution', 'Range', 'Power'])
    index = 0
    for sensor in device_json['device']['sensors']:
        sensor_type = safer(sensor, 'type')
        sensor_name = safer(sensor, 'name')
        sensor_vendor = safer(sensor, 'vendor')
        sensor_delay = safer(sensor, 'delay')
        sensor_resolution = real(safer(sensor, 'resolution'))
        sensor_range = real(safer(sensor, 'range'))
        sensor_power = real(safer(sensor, 'power'))
        sensors_sheet.append([index, sensor_type, sensor_name, sensor_vendor, sensor_delay, sensor_resolution, sensor_range, sensor_power])
        index += 1

def log(entries, log_sheet):
    log_sheet.title = 'Log'
    for entry in entries:
        log_sheet.append([entry])

def samples(device_json, data, samples_sheet, index):
    samples_sheet.title = 'Sensor %d' % index
    sensor_type = device_json['device']['sensors'][index]['type']
    entry_format = type_to_format[sensor_type]
    entry_size = struct.calcsize(entry_format)
    offset = 0
    while True:
        try:
            entry = struct.unpack_from(entry_format, data, offset)
            offset += entry_size
        except:
            break
        samples_sheet.append(entry)

def unzip(name):
    with zipfile.ZipFile(name, 'r') as zip_file:
        device_json = json.loads(zip_file.read('device.json'))
        entries = zip_file.read('runtime.log').split('\n')
        xlsx = openpyxl.Workbook()
        device(device_json, xlsx.active)
        sensors(device_json, xlsx.create_sheet())
        log(entries, xlsx.create_sheet())
        for index in range(100):
            try:
                data = zip_file.read('data.%02d.bin' % index)
            except:
                continue
            samples(device_json, data, xlsx.create_sheet(), index)
        xlsx.save(name.replace('.zip', '.xlsx'))

def main():
    if len(sys.argv) < 2:
        print('Please give a path to a ZIP file as a parameter')
        sys.exit(0)
    unzip(sys.argv[1])

main()
